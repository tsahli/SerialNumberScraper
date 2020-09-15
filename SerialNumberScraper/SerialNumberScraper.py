import tika
from tika import parser
import re
from itertools import tee, islice, chain
import openpyxl

def previous_and_next(some_iterable):
    prevs, items, nexts = tee(some_iterable, 3)
    prevs = chain([None], prevs)
    nexts = chain(islice(nexts, 1, None), [None])
    return zip(prevs, items, nexts)

plans = input("Enter the name of the PDF plans to scrape: ")
plans = plans + ".pdf"
raw = parser.from_file(plans)
content = raw['content']
splitContent = content.split('\n')
cleanList = []
deviceAndSerialList = []
r = re.compile("\S\S\S\S\S\S\S\S\s*")

for element in splitContent:
    if element:
        cleanList.append(element)

for previous, item, next in previous_and_next(cleanList):
    if (len(item) == 8 or "(" in item) and ("$MSD" not in item):
        deviceAndSerial = previous + " - " + item
        deviceAndSerialList.append(deviceAndSerial)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Sheet1'
nextRow = 1

for element in deviceAndSerialList:
    splitList = element.split(" - ")
    device = splitList[0]
    serial = splitList[1]
    sheet.cell(row = nextRow, column = 1, value = device)
    sheet.cell(row = nextRow, column = 2, value = serial)
    nextRow += 1

wb.save("serialNumbers.xlsx")